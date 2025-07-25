import pandas as pd
import requests
import json
import math
from openpyxl import load_workbook
from openpyxl.styles import Font

class ATS_BaseImplementation:
    INPUT_PATH = "D:/Automation/API_Automation/Input/baseImplementationATS.xls"
    CLIENT_ID = "47c6b8ee-0a96-4cd0-a282-a5990f97ce52"
    CLIENT_SECRET = "bYuG5yCzosD0DmB6Pn9vBDXXDNbaModh"
    TOKEN_URL = "https://amsin.hirepro.in/py/oauth2/160ba874540e42bb81f7192d29061a42/access_token/"

    def __init__(self):
        self.token = None
        self.input_data = []
        self.actual_data = []

    def get_token(self):
        header = {"content-type": "application/json"}
        data = {
            "client_id": self.CLIENT_ID,
            "client_secret": self.CLIENT_SECRET
        }
        request = requests.post(self.TOKEN_URL, headers=header, data=json.dumps(data), verify=True)
        response = request.json()
        self.token = response.get("access_token")
        print(f"Base Implementation Token: {self.token}")

    def get_data(self):
        df = pd.read_excel(self.INPUT_PATH, engine="xlrd")
        df = df.where(pd.notnull(df), None)
        self.input_data = df.to_dict('records')
        # print(self.input_data)

    def register_tag_candidate_to_test(self):
        for item in self.input_data:
            header = {"content-type": "application/json", "Authorization": "bearer " + self.token}
            if math.isnan(item["Test Id"]):
                test_id = 0  # or any default value
            else:
                test_id = int(item["Test Id"])
            # if math.isnan(item["ATS Candidate Id"]):
            #     ats_candidate_id = ""  # or any default value
            # else:
            #     ats_candidate_id = item["ATS Candidate Id"]
            data = {
                "atsCandidateId": f"{item["ATS Candidate Id"]}",
                "primaryEmail": f"{item["Email"]}",
                "firstName": f"{item["First Name"]}",
                "middleName": f"{item["Middle Name"]}",
                "lastName": f"{item["Last Name"]}",
                "gender": f"{item["Gender"]}",
                "dob": "05/02/1987",
                "profilePhoto": "/9j/4AAQSkZJRgABAQAAAQABAAD/2wCEAAkGBxMTEhUTExIVFRUVGBgXFxgXFRUVFxgXFRcWFxcXFxgYHSggGBolHRcXITEhJSkrLi4uFx8zODMtNygtLisBCgoKDg0OGhAQGi0lHyUtLS0rLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLf/AABEIAL8BCAMBIgACEQEDEQH/xAAcAAABBAMBAAAAAAAAAAAAAAAEAwUGBwABAgj/xABDEAABAwIDBQUECAQEBgMAAAABAAIDBBEFITEGEkFRYRMicYGRB6Gx8BQyQlJywdHhI2KC8TRzkrIIJCUzorMVF2P/xAAaAQADAQEBAQAAAAAAAAAAAAACAwQBBQAG/8QAJREAAgICAgICAgMBAAAAAAAAAAECEQMhEjEiQQRRMmEUcYET/9oADAMBAAIRAxEAPwC5CtELpy1ZEKs4a3ioDtt7SYqXeZDaR7Wuvn3Q4tG63Xm9hvnloOTz7RK5sNIXdq6F53uze02N2sdIQQciC1hFjz55ryzNOXEkkm+evFa9HoqwrFMYlnkdLK8uc4k9Bc3yHAfsgd/1XCyyAadErGvWNbdasvHjvfvqtjr6pJKR8QvHgqgqnxSNe0kFpuCDY26FXNshtzIY4XSu3x2rIb5bw3gBZ+Q6OB6KkW5/l+iecArNyWLfP8PfY544dw626AleQMlZ6pGaRkCSw+sErGvjex7XAZtN8jxREjUQoAmCEe+yPmCb5gsYaIO8aEDQ5pSAWJtzWO1sOSDr6zsmPf6DqVyJ7idGGp7G/aXGuzvGw986n7oP5qEuJOfvKWqJjI5znG+dz1KHe++mQTMWPijMk+TNuHNyRlNs969uhWOFvm64cD8k/knJC2zQqSNdPBEMeDoSFlLTiRpHEZjh8UM1hbl7uKNxQPJhduY8wlGnLmELHOR+6IjcDp3SlyVBpiwN/nNOGBYu6mkDtWHJw6fqmy/PIrkutqlOKeg1Jp2i4oZ2yMD2G7TmFsfFQHYvGezf2Lj3H/VvwdyVgBnFSSjxdF0J8lZtmi6k09FjQldwWXkeYjUvyWJGsaRmsRKddgOF9FoFae4AEk2AFyeQGq6K04ZLunBZ5u9rG3T65wpxGI4opC9puS5+RDXEEDdyJy6qvLKR+0U/9Sqha1pXA2N7nif26JjY3LRDJjorQkyNF/Q75+CJpqEnK3X593qpDSYBK5uTenifkBA5pBqDZHo6INBceH5/2Q8lHY+Iv7wptHgbu8N29j5XGp8Mz6IR+EOIJDb25i2nwF7e5apoxxZEfoh3ug/RK/QD6/nmpxh2ycsrbhhz6cwbBdzbG1bPsZHXXLM5+9ec4mqLK9fBYg9PelHMubacVKKnZ2QZlhv4fdufgU1Yhh5jaCePzms5pnnFoO2R2xmoJAWntIie9GTY+LTwPuPEcV6KwnE46mCOeM3ZI0OB456g8iDl5LydPfmCr19hUznUEjSbtbK7d/l3gC5vrZ39aYhU17J9KE3zBOMqAnC1gogUzbEZ2FlDdo64yP3Acm/HiVKNo6gRxk8eCgb/AHnVcfGdXIqYkWjQaBcualXWH6c1uOnv3n6Dgn3SFUDcL6Dmf0Q3Zuf0A5p4p6EynLJt7X4fupBhuyzpcmDu314eS8ppGODfRE4L2swa8eaVFDl3reqtCm2CDRoCeqRq9iTbWyyc5fQUcaKyNG373uKSdRH7LgVMa7Zp7cgN7wUfrKItNiCEEcrfs2UK9De17hk8G3PWyy9tc2niskeR83XMco5eI4Hw5J1ewBRvdIN8uBVn7N4n20Iue83I/kfnkqvI3Rl3o3ZdQeXipFsZiO5KGE3a7Q9D+h/NT542rQ7DLjKiyIzmlgkIxoiOSRFFbYLVG4IssS0w1WLasEsgrAsKwLunz7PHO0Id9LqN/N3bS7x5u7R1/enDAMFdLnc7vogNoZN+sqHfenlPrI4qxNjqYdgOeqnzz4rRbggpPY7YFs80briAfip1Q4a2w7uSb8GgyClVJHko43J7Kp1HoSgwiO1twZpF2y8J+zle5tlc9bap9Y1KBUpEjkBQUTWABrQAAuJob8E4EJORq80apEQxWhB4DJQnaHA2StLbWOdj16qz6+FRTEYrOPRTu4spVNHn/EISyR0bhYjL9FcfsEP/ACtQL5dsDbkTG3P55KrdtmD6S8jn8OCsH/h9qDasj4fwn36nfB+C6EHaTOfkXZbMqBmR0qBnRikVBtZN9RvK5KjYHHifhzTxtE/emdyH5JrbHvH5+bLj43o68tsymgublPmH4T2pF/qjQJupCHSBg4aqzdncNG63JG7bB0JYJs+0AAtBHgpph+HgDILujpQAnOBtlTixJCZ5Po5bTBJyUw5I0JNzVRQmxnqcPafshRDajZjtGFwAuNLKwJAhZFPkxJjYzZ51xClLSWuCaZaQagkeCvPbLZYVLS9gAkA/1dFS2I0j4nFrwQRwKTG4ug3vYNDdnJwIs4aX/ddtD4iJGZtGef5/qkt9H0swI3T8jitk2eiWbs7izaiBsjfBw4tcNQfj5p7YbqrNgagxVboA7uvByPNly0j3q0oTqppR4ypdFkJco2cT6La5eNc1iGwiyCsWFYu4cBnjerH/ADEvSR/+4qxtipt5gA4D46KDbZ0H0evqoeDZn7v4SS5vucFNtgYN2JpP2hveunuUvyV4l3xuyz8Lb3R1UipnJjw4ZABPtMxIhBobOQc0rsFaY1d7qoolbOXOSbylt1cuaF6j1jZVKLYvHqVLKxoso3i1t0ngkzgUQmUJtw0CoNuNr/D4Kc/8Po79b+GH4yKKe0WC07T95p9QQpn7AHNDKq5Ae5zLDiQxpvY8frKrF+KJc3bLUmQMyPlKb59U4nRReIzXefUpEybjMtSL+SSnN3uHzmc0nI69z4/kFx4x0kdZvs3sdLeU3NzdXhgOgXn/AGcn3KgdTZX1s9Wxhou9vqqMi8xUH4kwpWo5jEBQVsTsg9p8wnRhVMKrQid2aLg0XJt4qO1m0G+/sqZvav4kfVb1c7gEVi9E6buucWs6cU1vk+jNENNHvSOOQyu5x0Lj5E9AD5jKbukFGKqxzZgRcLzzOJP2WOLGjoLZlZJhMbfqOkH9bnD/AMiVV+OVuKCR7JqsQlu9ZjLDeItYNsLuBysciQU+7LUNa6c71c6WnaftR2e/LLO5sFrX6PJ37Jq2O3VRna3ZmOoBO6A/nz8VNBBkm+qbYpc462Mxy2UJi+y8kV+6cuQUYqWOYeRC9Jmia5wuAc1Xu12wt4amfd3XB73RtBveMHlwBGfmkq4v9DJJPorHD6hwmjlZk5rgfz/VXfRTh7GvH2gD65qjKDJ3S4I9VdWBNIgYDrutPqEvP+SGYOmEyalYu3tzv0W0hooTLGK082BPJbK5eMiPFdw4B5w9q+HSPc2uOZeTHLlbdcL7mQGlgRfoEthlY+GmiLGF7uzBtbLTIn91L9ssKdLTVgAvZrnW6su8Edcl1sNh4dG0EA2jYPcouTaVnTcVFvj1QxYNieKSRlzIww3zfI8AEcgP7ImDbmujO6Z6OVw+yJG3Fjb7wF/Ap6xvZ+SaZjXAimFt5gJ75B0dyaePNOdDs5uE7gDYnOJLLDRxaSDwy3QB097ItMTNNfs1gu3krrCeHc00vYjmDmFOsPrWysD2nIqNT4e27nBsbGm92taQLniM7A3zyRezZLQW8ECb5UFKEXG0H43ihhaLDMkfEX9yrjHNqa+7nRDcjB+sRvOOdrhoByzHDzU4xNu9K3ezA80jPQgxvjdu/wAQC7g03u03B15jTReblejFGNFUT43VPP8AErJGX+/BI1gubC7g2wGmZ5o6TDsQaWu7Zr2HiwmxtwPADyVgtwcgOa53aBzHMzBIAebuyJN7n05JDAsCNMws33OZe7Q7Mt6X4hG3SBiioPaTESIHkHe7zSBzNsslYPs8wj6OyFhA3gwlx/ncQXX5nvW8k37dUQdJBlpUQH1la0+4qbUUIE2XDj5HP3BbHdILpSb+hykQE5RsqAnKqZzzz/I/vHh3j7kjWu3R5fElHY5BuVErbcSR/UcvimvFXXHzyXKgto6stWNcEm65rjfIg5aq7oosNipmTSvdulocCHuzyvo1Ui1twpTsQHySNYCCYj20THAOaXNcCWm/C5vZVTSexMXROqTaKlY8iOnnbbO794EC17kcBbNWJs7jJeWtcCN4Bzb8nC7TnnYhRf8A+PdVv7WeMM3gN9rHOAeQLAu8hbI58VI6DC2RC7GNbpoADYaC/ReUEncTzcq8iQ1zkxuwz+K2Vr3NLb5DLM6m/hYWTy7MAruJgKbxVi1Jx6Gh+DiV+/IS42sC4k2GtvDonilowwWAt0CIa0BdhFRnJs4emiuKdnpuqWXScmxuPQE11hc+KiXtA2+pYIJIYpGy1DwYwwZ9ncWLn8rX01JUtnFgvOG3MNsTqf8AMv6gFBFema2b2Wou1mjYRkXXP4RmVc8Wo8FB/Z9hBa3t3CxcLMH8vPz/ACU1iKjySuRXiVRCPzWJOK+YWLENLHK0sKxdk4BDsYhAfO0mwkaWnwe0j4FRr2a1F2BpOYa1p8WjdPvCm2OxDtASMi0+oy+BCrnYR/Z1NTGBYNmlsDwBeXD4qOapnSg+UV/RbEcQKUMQSFPJcBLNctsVxYBiIACQwl4BKUxl1m3Gt7BC4NGSblBexq/ELrXDfBTjStDm6JmxlpBBGqc8KfdgKJdgSWgzsggawWujnvTXXuyK1gxRXO2kh7antmTPFYfhkafyU1wqM7znE3t3fn54qD4vTunxCmY3Ps3doQOO7n8bKxqSHcYAddT4nNNwx3YGedKjUqb504SJvqgqCMpXbiO09/vBp9L/AKKK1jrk9Ap5tlhckhEkY3iBYtGuV8xzUHfTm7rgg9RbVcrC1VnXzJpsGiZknXZCt7GqiedA+zvwuyPxv5IQMyCRaO8n3dk/R6kwuEWunB0WSgXst2mFRB2T3fxYQGu5ub9l/wCR6hWExOg1Rk1syJl2BIO3mnJExzNaLHVDl0hNxYN5WuT1vfIL0mjIhdNLvBLoWmFgiAVt6M9mnhByNRrzkhihZqY04mLC3NVJtJsyZsUkccmWY53XIC3uVv1g3nAcs1Cdqrx1Ub/syN3T4tJI9xUmZtbQ/Ek2rOoWgEACwAsB4JVuRKRJ0KWcbqYsFAM7rFxG/JYiR4sgrFhWLsnBBqykbILOGmYINiPAqqa2kFPiszWCzTuPte5JcwXOfUH1Vvqsts4CzEu04PhZb+klp8OCTnXjZR8aT5pEvw+a4RzSovSVobu56p7ZVWaT8FNF2VSRvFIy4AtsS03z8LJqpJ5o3udlIw/dBa5p6tJNx1CXkq5JDutaQM7nj5Dj7lqgw+RveNjfUDX1RcbMvQnPPNI8Oya0aNIJcT0FxYdSnvDGFrM9SScuCaq7DpHWcLDvaXtlz/slHVUrLXaCNMjfz/ZbxMb0PEjk2Yk7IoqOpDm9c/cmypk3r9MljPIj2BQ/9Sv/APk8+9vvzCm7goxs9EPpjjxEJHPV7ePkpTIFXh/Ahzu5gkoTdWHIpylTfWaFNF+iGVDcweSiGNMaZXADkpra+SYcXw8X3xyzXAiju5fshVfT7py5XTU9vuzUlxSPK/JR2oyN1ThlZNkjR3heNy0lQyeI95h04OafrNd0P6Fektk9pYq2Bs8RyOTmn6zHDVruo9+RXlqbUp72I2qkw+oD23dE+wlZ95vMfzDUenFVpaEt0em6iYDMpB2NMGQzTRUStrKdktPJdrhvAjj0N9D0QdJhYI75dvdSUFtspxY4NXJjrPtOAbAXPIZorAsQnmc57gGxjIC2ZPO6SocHbo1thxNrKQ09OGNDQMgi4t+zMrxpVFbFHPyQk0lglpTYIRrd49B7ysk6JkcRRHNx1KjW2FHvx34sIcPz9ymD2poxGG4I5pM42hkJbILT1NwCfBHMdqm2qp+ze5vA5jyRjHZjPgpEXpi8axJtesWoJFolaWytLtHzxije2uDiWLtW5SQBzmnm213sPja/iApIh68fwpPwO/2lY1aphRbTTRUJxcFjXNJsCC3PPk7z6Kc0jXOYN3O41vplqql2npnwWe24hfmbfYff3B1/UKa7H41/Ca2+80ADK1zYDnoFz4qjpSdkgfSz73/dG70bY+tzf0RENLLwlseosfcnGms4Xaf7oHE3yxAu3d4cLa8P3TU0eUzp9A8/WlJPQD90kMKzuXyW5b5F/RawqSaUgnut45Z5jTy0TwWdmLk+ul0TYMpsGip+xYc8gCRzz+SmU1I3X3yuTe/AePFdVuJuc4tOQc0i1jqNSDw8+iiUFUal5gYLsbnK4g3sc2tbnrz0sgYCZOdkqfuOmNv4h7v4G3DfXM+aeJELgP8Ah4wPsgt/0khFSq2C0jnzdydgcqa692RTpKmyuGRTQSKSusboPE7bt0a6MuHIdU24litPELOeHO+6O8brgxR9BOmiL4m3IqK14sCn/EMXMr7boa3Ow4+fJR+t1KdhVSJp9AJbkUk4Ihjbg9Sk5hmq0xLJTsPtZLQuBF3wuPfjv/5M5O+PvHojZ7EIqiNskZDmuFwfyPIjkvLWGtu1w5ZqabE7Uy0DzYb8TiN5l7WOm83kfiluajLZqi2j0dExbeo/hm0rJmhzQ4XHG36o/ty7VM/6Ra0L4P2ZO7eNhpxKUibyXG4lYwl9u2F0jHIOojujSEk9q1qzEyHY9h1zvAZhMbCRYdVPauC/BRzE8MN95vDUc1JODTsrx5PTGuRbSTzY2KxLRSWyVpbK0u0fPmJKs/7b/wALvgUqhcRqWMjeXuDRunXqOHNeZq7IBJQNlhLHjea5uYUDEb8Pn7Ml24TvRO3iAQL91xGd8x4qzqCO7AOgSON7Px1Mbo5ADfQ2za4aOHIqGOjoy2B7O7SEAM+sSQ0OLSLki5NvT105zKmkDwHOPXM2IsqSjnmoJQyVhLGE7r2i4IIsSfLW/wCSkcHtDiAbeQNvqRx8b/OSYv0Los4ndzbp488xkUy4pirgSMybb1gbd0WvkdfLmor/APYdPm4yC9rAC2fWyZq3FZqskU8JcDkJHAtaN63ezOZ42CLQDRziWLPkf2ERc6V5IFwAGtORcRa2Vs8+WeammA4MKeEMBuRm532nOOpdzQeyGzbYQXOO/IfrOPPjYcApc+LL8kD2EtHOAv7jm/defeAf1R8ii4qJIpTuEZi5adDb4JxkxxrWFz2PFtd1pf6BuZ9FRjyKqJcuOV2guZNVacl1R7QUs53Yp2PcNW3s8dC05gpDETkVQhLKyqqGtf3pibfd3wwDyamSronsvkB4fq7Mq4a2nPEAqFY7YAkgC1+XySuM4NHatMr1v19NNeJQkpu4joUVUSgE8yblBb1pG9cvXJMghbNU7Mh4/AIeVungE4Fm7YePxKDOd/MJkXsBi+DZuI53+Ce4o7sd4j4hMmDGz/T9FJ6SmLm7o4uHuU3yH5jca8SxtjKnuBTuldkoHspQOY0KdUwyR4rByILY+6IAQELu8j7qhCGY5c7qxzloyo1EEQmjQ0kCNe4IeWQWWOKCTIvj9ExrXSGzdwEknIWGt1tQT2zbUZCjjOvelI5fZZ56nwHNYg/jxew18iUdIvcoDFsXgpmdpUTMibze4C/QDUnoFBvaP7UoqLegpt2Wp0PGOL8VtXfy+q8/45jc9XIZaiV0jzxccgOTRo0dArKIlGy0toPa1U1c4psP/gxueGCQi8jrm28AcmDpYnwUmjwcxxOc975Hkd573FznHnc6DpoqEwer7KVkg1jc1467hBsvTtaWyQB7M2vYHDwc24SPkXxVFGFJNnGFDujwTl2OWSBwxvdCdokhdDWyI4xBaS5GuvHzReG4fE7Ixxm2ncH6J0xmj3he2iApoyMxkeXAoa2FdoNjwmFufYs8mt4eXL4oapaOGQ4Ae6yN7ckW4rgU5vYcePFMSF79iuE0+Xz8lOMrF3TQ2aF08LQWyN4hFaRp629U7x0YcL2SGIwX8vyTlRHuhDWzeWilfbjgbYDBVx9x5cWOc3I3tvNdccciLqK4L7R6qMbkx7dlrd7J4/q4+anX/ELiLezp4BqXuf4Bo3fi73KkFdibUUJnFPs9E7Q7RRRDMk68zYjgbaeaqnH8d7U2uCNTa/opTt49riS4Ny4jIuJ5KspXDeNuHVcyHm9l0taRuok6rKkXAcEjNqUTRneBafJPapWAGTPvuu57v6fFARnvEIhzSGEHhp8f1Q1+9fzQwWjH2E0DbOI638tVPNk4t5wKhNPHdwtxyVibKtDWj50/spsvlJDsekWJhMFgnF826E34e+4CcNy6rhHQmb2bpnkZ8TojWRO4lagiCLATFEW2Bva5IGN/BOwYtkBFxA5jUKZ51KHxFoiifLK6zI2ue7waCT8E+OcAqs9uG0W5TNpmmxmPe/AyxPqd0ea3ijzmykcXrXTzSSu1kcXHpfQeQsPJYg7rEQFhe0P+KqP86X/2OTcQnHaD/FVH+dL/AL3ICydWjTInWKvP2U7QielNI89+Edz+aInL/STbwsqLITlgGMyUszJozZzDcciOLT0OiCULjxf+GXTtHqGhisAjw1NeyWMxVsDZojr9ZvFjuLSnt0ajcWtDeVibm3CboowDYhOkHJcT0980NGp+hJsTenqu4wCckkIUVTNWo9IXAWWXRau2tRUAAyFtyC4crXGqZca2ghpY3PfIGtaLnj4ADiTyTb7Q9taWhaWlrZJ3Duxi18+Lz9lvx4LzzjmPz1Upkkfnwa3usb0aPz1TY472ZYXtlj0tdUunc1zW/VY3PusF7X6nMnxTE5ttcvFddu77x9StOeTqSfFUxVaBJftLiJlc5+gzDQoyOIRUslwQUHZc3FGkWN2zqqGd+i1TGxvyCWc3eFiEiMgUxPVA+x1hIlb14/shfozm5EZDj0/IreHHIg+SKFY8WuA4aX4pO4tpBafZukOnQqf4M6zWjkFCKbddmMlLsJc8gAMJ63H6pXJOQyK0WPg/1Qn+JqYMFuGjesCn2J6uh0Tz7DmLveQoetlyMVQX2qTdMhjIuXLx6jdTU2BzXmz2jY19JrZCDdkf8Nv9J7x83X9Ari9omOfRqSRwPfI3WfidkD5a+S86kokBIxrbrE9bG4V9Jq4orZOdd34GZu+FvNYhcqNjGwDaD/FVH+dL/vcm9H7Qf4qo/wA6X/e5N5VSejDFi0sCC7NJDsbtdPh8u/Ed5rrb8ZPdcB8D1Xo3Y3bSlxCO8TrPA78TsntPhxHUZLykl6CukhkbLE90cjTdrmmxH7dEEoWePYjm2N0u5t1U3s49qH0pwp6oATfZe0HdktzA+q73HorXY/JTuNBWJugWRtsUQHiyFa/Moao0KJsq59pftMjogYIC19SRpq2O/F/M/wAvqo57TPasWufS0RLXAlskxFiCMi2MHj/MfLmqWfISSSSSTckm5JOpJ4lUQx/YAtW1j5Xukke573m7nE3JJ5pEtstjLxXN1QkYaK2FhWLaPBMkma0yU/Oa4cVkKgrRRewiKU5C3oiHw72hz5IGDiu2OPogcd6NTD6emcLZaIv6OTw+eCbmTua4Wdll708U8t/BIyclsZGmdYVSEb91Pdn47huXD3qLUthG93QnztYKf7OU1o2X13R7wEuD5SsZVIklFHonSNqBpwLIoSK5E0ggFYSh+2WCW62waCQ5Jzy2C43lG9u8e+iUz5Bm76rB/O7IX6cfJEjGVZ7Wcd7apELT3Ydbffdr6Cw8yoIu5pC4lxNySSSdSSSST5rlqMS3Za3sNwu75pyPqhsbT1Pef8GrSmHseowzDmO4yPe8/wCotHuaFiCxq6P/2Q==",
                "thirdPartyCandidateJson": {
                    "key": "value"
                },
                "fileName": "qwe.jpeg",
                "phoneNumber": f"{item["Mobile Number"]}",
                "testId": str(test_id)
            }
            request = requests.post("https://amsin.hirepro.in/py/ats/v1/registerAndTagCandidateToTest/",
                                    headers=header, data=json.dumps(data), verify=True)
            response = request.json()
            if 'error' in response and response['error']:
                response_msg = response['error'].get('errorDescription', 'No error description provided')
                item["Actual"] = response_msg
            elif 'data' in response and response['data']:
                response_msg = response['data'].get('message', 'No message provided')

                item["Actual"] = response_msg
            else:
                response_msg = 'No relevant information found'
                item["Actual"] = response_msg
            print(f"Response message - {response_msg}")
            self.actual_data.append(item)
        for row in self.actual_data:
            row['status'] = 'Pass' if row['Actual'] == row['Expected'] else 'Fail'

            # Move 'status' to first column
        processed_data = []
        for row in self.actual_data:
            new_row = {'status': row.pop('status')}
            new_row.update(row)
            processed_data.append(new_row)

        # Convert to DataFrame and write to Excel
        df = pd.DataFrame(processed_data)
        output_file = "D:/Automation/API_Automation/Output/baseImplementationATS.xlsx"
        df.to_excel(output_file, index=False)

        # --- Begin formatting logic ---
        wb = load_workbook(output_file)
        ws = wb.active

        # Find column indices
        headers = [cell.value for cell in ws[1]]
        status_col = headers.index("status") + 1
        actual_col = headers.index("Actual") + 1
        expected_col = headers.index("Expected") + 1

        # Apply color formatting row by row
        for r in range(2, ws.max_row + 1):
            actual = ws.cell(row=r, column=actual_col).value
            expected = ws.cell(row=r, column=expected_col).value
            match = actual == expected

            font = Font(color="008000" if match else "FF0000", bold=True)
            font1 = Font(color="008000" if match else "FF0000", bold=False)

            # Apply to status, actual, expected
            ws.cell(row=r, column=status_col).font = font
            ws.cell(row=r, column=actual_col).font = font1
            ws.cell(row=r, column=expected_col).font = font1

        wb.save(output_file)

if __name__ == "__main__":
    BI_ATS = ATS_BaseImplementation()
    BI_ATS.get_token()
    BI_ATS.get_data()
    BI_ATS.register_tag_candidate_to_test()
