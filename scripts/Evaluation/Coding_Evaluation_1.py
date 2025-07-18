import json
import time
import pandas as pd
import mysql
import mysql.connector
import math
import numpy as np
import requests
from openpyxl.styles import PatternFill
import os


class CodingQpEvaluation:

    def __init__(self):
        print("__init__")
        self.input_data = None
        self.actual_data = None
        self.appName = "crpo"
        self.crpo_tkn = ""
        self.input_path = r"D:\Automation\API_Automation\Input\codingQP_Evaluation.xls"
        self.output_path = r"D:\Automation\API_Automation\Output\codingQP_Evaluation.xls"

    def crpo_login(self):
        crpo_login_header = {"content-type": "application/json", "APP-NAME": self.appName, "X-APPLMA": "true"}
        crpo_login_payload = {"LoginName": "admin", "Password": "4LWS-0671", "TenantAlias": "automation",
                              "UserName": "admin"}
        crpo_login_request = requests.post('https://amsin.hirepro.in/py/common/user/login_user/',
                                           headers=crpo_login_header,
                                           data=json.dumps(crpo_login_payload), verify=True)
        crpo_login_response = crpo_login_request.json()
        self.crpo_tkn = crpo_login_response.get("Token")


    def fetch_input_expected(self):
        df = pd.read_excel(self.input_path, engine="xlrd")
        df = df.where(pd.notnull(df), None)
        self.input_data = df.to_dict('records')
        for item in self.input_data:
            for key, value in item.items():
                if isinstance(value, float) and np.isnan(value):  # Check if the value is NaN
                    item[key] = None  # Replace NaN with None


    def attend_evaluate(self):
        for item in self.input_data:
            login_name = item['Login_Name']
            password = item['Password']
            test_login_header = {"content-type": "application/json", "APP-NAME": self.appName, "X-APPLMA": "true"}
            test_login_payload = {"ClientSystemInfo": "Browser:chrome/60.0.3112.78,OS:Linux x86_64,IPAddress:10.0.3.83",
                                  "IPAddress": "10.0.3.83", "IsOnlinePreview": False, "LoginName": login_name,
                                  "Password": password, "TenantAlias": "automation"}
            test_login_request = requests.post(
                "https://amsin.hirepro.in/py/assessment/htmltest/api/v2/login_to_test/", headers=test_login_header,
                data=json.dumps(test_login_payload), verify=True)
            test_login_response = test_login_request.json()
            test_login_tkn = test_login_response.get("Token")
            print("Test login token : ", test_login_tkn)

            submit_test_header = {"content-type": "application/json", "APP-NAME": self.appName,
                                  "X-APPLMA": "true",
                                  "X-AUTH-TOKEN": test_login_tkn}
            submit_test_payload = {"disableBlockUI": False, "isPartialSubmission": False, "totalTimeSpent": 27,
                                   "testResultCollection": [
                                       {"q": item["Q1_Id"], "timeSpent": 8, "timeSpentOnTicker": 0,
                                        "secId": item["Section_1_Id"],
                                        "a": item["Q1_Code"], "l": item["Q1_Lang_Id"]},
                                       {"q": item["Q2_Id"], "timeSpent": 19, "timeSpentOnTicker": 0,
                                        "secId": item["Section_2_Id"],
                                        "a": item["Q2_Code"], "l": item["Q2_Lang_Id"]},
                                       {"q": item["Q3_Id"], "timeSpent": 27, "timeSpentOnTicker": 0,
                                        "secId": item["Section_3_Id"],
                                        "a": item["Q3_Code"], "l": item["Q3_Lang_Id"]},
                                       {"q": item["Q4_Id"], "timeSpent": 48, "timeSpentOnTicker": 0,
                                        "secId": item["Section_3_Id"],
                                        "a": item["Q4_Code"], "l": item["Q4_Lang_Id"]},
                                       {"q": item["Q5_Id"], "timeSpent": 53, "timeSpentOnTicker": 0,
                                        "secId": item["Section_4_Id"],
                                        "a": item["Q5_Code"], "l": item["Q5_Lang_Id"]},
                                       {"q": item["Q6_Id"], "timeSpent": 58, "timeSpentOnTicker": 0,
                                        "secId": item["Section_4_Id"],
                                        "a": item["Q6_Code"], "l": item["Q6_Lang_Id"]}]}
            submit_test_request = requests.post(
                "https://amsin.hirepro.in/py/assessment/htmltest/api/v1/finalSubmitTestResult/",
                headers=submit_test_header,
                data=json.dumps(submit_test_payload, default=str), verify=True)
            submit_test_response = submit_test_request.json()
            print("Submit Test Token : ", submit_test_response.get("systemTkn"))

    def evaluate(self):
        eval_assessment_header = {"content-type": "application/json",
                                  "X-AUTH-TOKEN": self.crpo_tkn,
                                  "APP-NAME": self.appName, "X-APPLMA": "true"}
        eval_assessment_payload = {"testId": 8581}

        eval_assessment_request = requests.post(
            "https://amsin.hirepro.in/py/assessment/eval/api/v1/eval-online-assessment/",
            headers=eval_assessment_header,
            data=json.dumps(eval_assessment_payload, default=str), verify=True)
        eval_assessment_response = eval_assessment_request.json()
        GUID = eval_assessment_request.headers['X-GUID']

        print(GUID)

    def fetch_actual(self):
        time.sleep(300)
        cursor = None
        conn = None
        question_ids = None
        self.actual_data = self.input_data
        for item in self.actual_data:
            question_ids = [item["Q1_Id"], item["Q2_Id"], item["Q3_Id"], item["Q4_Id"], item["Q5_Id"], item["Q6_Id"]]

        try:
            conn = mysql.connector.connect(host='35.154.36.218',
                                           database='appserver_core',
                                           user='qauser',
                                           password='qauser')
            cursor = conn.cursor()
        except Exception as e:
            print(e)
        testcase_marks_query = f"SELECT question_id, count(id) FROM coding_question_attachments WHERE question_id IN ({', '.join(map(str, question_ids))}) AND is_sample=0 GROUP BY question_id;"
        df = pd.read_sql_query(testcase_marks_query, conn)
        qustionwise_testcase_count = df.set_index('question_id')['count(id)'].to_dict()

        for item in self.actual_data:
            cursor.execute(
                "select question_id, obtained_marks from test_results where testuser_id = %d;" % item["Test_User_Id"])
            data = cursor.fetchall()

            # Create a dictionary for quick lookup
            question_marks_dict = {i[0]: i[1] for i in data}

            # Iterate over Q1_Id to Q6_Id and update the corresponding marks
            for i in range(1, len(data)+1):
                q_id = item.get(f"Q{i}_Id")
                if q_id in question_marks_dict:
                    item[f"Actual_Q{i}_Marks"] = question_marks_dict[q_id]
                else:
                    print(f"Question Id for Q{i} not found in the actual data")
            count = 1
            for qid in question_ids:
                cursor.execute("select tri.coding_question_attachment_id testcase_id ,tri.coding_obtained_mark marks "
                               "from test_result_infos tri inner join test_results tr on tr.id = tri.testresult_id inner join test_users tu on tu.id = tr.testuser_id "
                               "where tu.test_id=%d" % item["Test_Id"] + " and tr.testuser_id = %d" % item["Test_User_Id"] + " and tr.question_id in (%s)" % qid + ";")
                test_results_question_tcwise_marks = cursor.fetchall()
                question_tc_marks_dict = {i[0]: i[1] for i in test_results_question_tcwise_marks}
                len_question_tc_marks_dict = len(question_tc_marks_dict)
                len_qustionwise_testcase_count = qustionwise_testcase_count.get(qid)
                if len_qustionwise_testcase_count == len_question_tc_marks_dict:
                    for i, (_, value) in enumerate(question_tc_marks_dict.items(), start=1):
                        item[f"Actual_Q{count}TC{i}_Marks"] = value
                else:
                    if test_results_question_tcwise_marks is None:
                        for i in range(0, len_qustionwise_testcase_count):
                            item[f"Actual_Q{count}TC{i+1}_Marks"] = "Empty"
                    elif test_results_question_tcwise_marks is not None:
                        remaining_tc_column = len_qustionwise_testcase_count - len_question_tc_marks_dict
                        flag = 0
                        for i, (_, value) in enumerate(question_tc_marks_dict.items(), start=1):
                            item[f"Actual_Q{count}TC{i}_Marks"] = value
                            flag += 1
                        for i in range(0, remaining_tc_column):
                            tc_index = flag + 1
                            item[f"Actual_Q{count}TC{tc_index}_Marks"] = "Empty"
                            flag += 1
                count += 1

            view_candidate_score_by_candidate_id_header = {"content-type": "application/json", "APP-NAME": self.appName,
                                                           "X-APPLMA": "true",
                                                           "X-AUTH-TOKEN": self.crpo_tkn}
            view_candidate_score_by_candidate_id_data = {"testId": item["Test_Id"], "candidateId": item["Candidate_Id"],
                                                         "reportFlags": {'testUsersScoreRequired': True,
                                                                         'fileContentRequired': False}, "print": False}

            # print(view_candidate_score_by_candidate_id_data)
            view_candidate_score_by_candidate_id_request = requests.post(
                "https://amsin.hirepro.in/py/assessment/report/api/v1/candidatetranscript/",
                headers=view_candidate_score_by_candidate_id_header,
                data=json.dumps(view_candidate_score_by_candidate_id_data, default=str), verify=True)
            transcript_response = json.loads(view_candidate_score_by_candidate_id_request.content)
            grp_sec_marks = transcript_response['data']['groupAndSectionWiseMarks']
            test_total_marks = transcript_response['data']['assessment']['marksObtained']
            item["Actual_Test_Total"] = test_total_marks
            grp_count = 1
            sec_count = 1
            for grp in grp_sec_marks:
                grp_total_marks = grp['obtainedMarks']
                item[f'Actual_Grp_{grp_count}_Total'] = grp_total_marks
                # print(f'Actual_Grp_{grp_count}_Total : ', item[f'Actual_Grp_{grp_count}_Total'])
                for sec in grp['sectionInfo']:
                    sec_total_marks = sec['obtainedMarks']
                    item[f'Actual_Sec_{sec_count}_Total'] = sec_total_marks
                    # print(f'Actual_Sec_{sec_count}_Total : ', item[f'Actual_Sec_{sec_count}_Total'])
                    sec_count += 1
                grp_count += 1

    def compare_write(self):
        df = pd.DataFrame(self.actual_data)

        # Create an Excel writer
        with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Comparison")

            # Access the workbook and sheet
            workbook = writer.book
            sheet = workbook["Comparison"]

            # Define color fills
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            # Loop through the columns dynamically
            for col in range(1, len(df.columns) + 1):  # Iterate through all columns
                if 'Expected' in df.columns[col - 1]:
                    # Find the corresponding Actual column by replacing 'Expected' with 'Actual'
                    expected_column = col
                    actual_column = col + 1  # Assuming Actual column follows the Expected column

                    # Check if actual column exists, and compare
                    if actual_column <= len(df.columns):
                        # Start from row 2 (skip header)
                        for row in range(2, len(df) + 2):
                            expected_value = sheet.cell(row=row, column=expected_column).value
                            actual_value = sheet.cell(row=row, column=actual_column).value

                            # Apply coloring based on comparison
                            if expected_value == actual_value:
                                sheet.cell(row=row, column=actual_column).fill = green_fill
                            else:
                                sheet.cell(row=row, column=actual_column).fill = red_fill

        # Save the file (automatically done with ExcelWriter context)
        print("Excel file 'comparison_output_dynamic.xlsx' has been created with color-coded cells.")

if __name__ == "__main__":
    cqpe = CodingQpEvaluation()
    cqpe.crpo_login()
    cqpe.fetch_input_expected()
    cqpe.attend_evaluate()
    cqpe.evaluate()
    cqpe.fetch_actual()
    cqpe.compare_write()
